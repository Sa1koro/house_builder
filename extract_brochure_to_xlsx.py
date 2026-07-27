#!/usr/bin/env python3
"""圣都整装长图宣传页 → 结构化 Excel（AEs vs A5s 硬装全案）。

用法:
  .venv/bin/python extract_brochure_to_xlsx.py \\
      --aes 8B64F6AC025B21F9083E98A9C885A690.PNG \\
      --a5s 2C5AB46DD246F16FC0B2261366FE1E8A.PNG \\
      --area 76.34 --sales-area 90

依赖: openpyxl, pillow；可选 tesseract（chi_sim）用于辅助校验价表。
价表与配置以脚本内已校验的结构化数据为准（来自宣传页 OCR+人工核对）。
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
HEADER_FILL = PatternFill("solid", fgColor="1F4E3D")
HEADER_FONT = Font(bold=True, color="FFFFFF")
AES_FILL = PatternFill("solid", fgColor="E8F5E9")
A5S_FILL = PatternFill("solid", fgColor="E3F2FD")
DIFF_FILL = PatternFill("solid", fgColor="FFF8E1")
WARN_FILL = PatternFill("solid", fgColor="FFEBEE")


def style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN


def autosize(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def calc_quote(area: float) -> dict:
    """50 ≤ area < 80 档公式（宣传页）。"""
    if not (50 <= area < 80):
        raise ValueError(f"当前脚本默认公式覆盖 50≤面积<80，收到 {area}")

    aes_hard = 65000 + (area - 50) * 699
    a5s_hard = 75800 + (area - 50) * 799
    custom_p, custom_s = 13000, 16000

    aes_mgmt, a5s_mgmt = aes_hard * 0.12, a5s_hard * 0.12
    aes_pm, a5s_pm = aes_hard * 0.02, a5s_hard * 0.02

    return {
        "aes_hard": aes_hard,
        "a5s_hard": a5s_hard,
        "custom_p": custom_p,
        "custom_s": custom_s,
        "aes_base_p": aes_hard + custom_p,
        "aes_base_s": aes_hard + custom_s,
        "a5s_base_p": a5s_hard + custom_p,
        "a5s_base_s": a5s_hard + custom_s,
        "aes_mgmt": aes_mgmt,
        "a5s_mgmt": a5s_mgmt,
        "aes_pm": aes_pm,
        "a5s_pm": a5s_pm,
        "aes_total_p": aes_hard + custom_p + aes_mgmt + aes_pm,
        "aes_total_s": aes_hard + custom_s + aes_mgmt + aes_pm,
        "a5s_total_p": a5s_hard + custom_p + a5s_mgmt + a5s_pm,
        "a5s_total_s": a5s_hard + custom_s + a5s_mgmt + a5s_pm,
    }


AES_PRICE_TABLE = [
    (50, 78000, 81000),
    (60, 84990, 87990),
    (70, 91980, 94980),
    (80, 98000, 101000),
    (90, 104990, 107990),
    (100, 111980, 114980),
]

A5S_PRICE_TABLE = [
    (50, 88800, 91800),
    (60, 96790, 99790),
    (70, 104780, 107780),
    (80, 110500, 113500),
    (90, 118490, 121490),
    (100, 126480, 129480),
    (110, 147500, 150500),
    (120, 155490, 158490),
    (130, 163480, 166480),
    (140, 182120, 185120),
    (150, 194200, 197200),
    (160, 206280, 209280),
    (170, 218360, 221360),
    (180, 230440, 233440),
]

CONFIGS = [
    ("客餐厅", "地砖+踢脚线", "马可波罗/东鹏/欧神诺/蒙娜丽莎", "马可波罗/东鹏/欧神诺/蒙娜丽莎", "相同"),
    ("客餐厅", "腻子基层", "东方雨虹", "东方雨虹", "相同"),
    ("客餐厅", "乳胶漆", "多乐士/嘉宝莉", "多乐士/嘉宝莉", "相同"),
    ("客餐厅", "入户/门套", "升升概念/江山欧派/派的门", "TATA/升升概念/江山欧派/派的门", "A5s优势：可选TATA"),
    ("客餐厅", "窗台石", "天然石/人造石/岩板", "天然石/人造石/岩板", "相同"),
    ("客餐厅", "强电电线", "中大元通/东方电缆", "飞利浦/中大元通/东方电缆", "A5s优势：可选飞利浦"),
    ("客餐厅", "弱电电线", "永鼎", "飞利浦/永鼎", "A5s略优"),
    ("客餐厅", "线管底盒", "伟星/日丰/公元", "伟星/日丰/公元", "相同"),
    ("客餐厅", "开关插座", "西门子", "西门子", "相同"),
    ("客餐厅", "全屋定制品牌", "志邦/金牌/莫干山/兔宝宝", "志邦/金牌/莫干山/兔宝宝", "相同"),
    ("厨房", "地砖墙砖", "马可波罗/东鹏/欧神诺/蒙娜丽莎", "马可波罗/东鹏/欧神诺/蒙娜丽莎", "相同"),
    ("厨房", "集成吊顶+电器", "奥普/美的", "奥普/美的", "相同"),
    ("厨房", "橱柜", "金牌（约3m地柜+1.5m吊柜）", "志邦/金牌/莫干山/兔宝宝", "A5s优势：品牌可选更多"),
    ("厨房", "台面", "厨之宝/欧铂利", "厨之宝/欧铂利", "相同"),
    ("厨房", "水槽+龙头", "欧琳/诺帝玛", "欧琳/诺帝玛", "相同"),
    ("厨房", "金属移门", "卡帝/德诺克", "卡帝/德诺克(梵)", "基本相同"),
    ("厨房", "门套", "升升概念/江山欧派/派的门", "TATA/升升概念/江山欧派/派的门", "A5s优势：可选TATA"),
    ("卧室", "复合地板+踢脚线", "德尔/莫干山/书香门地", "德尔/莫干山/书香门地/大自然", "A5s优势：多大自然"),
    ("卧室", "木开门+五金", "升升概念/派的门/江山欧派", "TATA/升升概念/江山欧派/派的门", "A5s优势：可选TATA"),
    ("卧室", "腻子/乳胶漆", "东方雨虹 + 多乐士/嘉宝莉", "东方雨虹 + 多乐士/嘉宝莉", "相同"),
    ("阳台", "地砖+踢脚线", "马可波罗/东鹏/欧神诺/蒙娜丽莎", "马可波罗/东鹏/欧神诺/蒙娜丽莎", "相同"),
    ("阳台", "专业防水", "东方雨虹", "西卡/东方雨虹", "A5s优势：可选西卡"),
    ("阳台", "地漏", "潜水艇", "潜水艇", "相同"),
    ("卫生间", "地砖墙砖", "马可波罗/东鹏/欧神诺/蒙娜丽莎", "马可波罗/东鹏/欧神诺/蒙娜丽莎", "相同"),
    ("卫生间", "专业防水", "东方雨虹", "西卡/东方雨虹", "A5s优势：可选西卡"),
    ("卫生间", "集成吊顶+电器", "奥普/美的", "奥普/美的", "相同"),
    ("卫生间", "浴室柜/龙头/马桶/花洒", "九牧/箭牌/惠达", "九牧/箭牌/惠达", "相同"),
    ("卫生间", "淋浴房", "宣传页本段未明确单列", "建霖智家/37°C2（约≤3.6㎡）", "A5s优势：明确含淋浴房"),
    ("卫生间", "木门", "升升概念/江山欧派/派的门", "TATA等（同卧室门品牌池）", "A5s优势：可选TATA"),
    ("辅材", "给水管PPR", "伟星/日丰/公元", "伟星/日丰/公元", "相同"),
    ("辅材", "防水/腻子/石膏", "东方雨虹", "东方雨虹", "相同"),
    ("其它工程", "垃圾袋/成品保护/竣工保洁", "含", "含", "相同"),
    ("售后", "服务承诺", "十诺 / 14815 / 隐蔽约5年", "同体系服务承诺", "基本相同"),
]


def build_workbook(area: float, sales_area: float) -> Workbook:
    q = calc_quote(area)
    wb = Workbook()

    # --- Sheet 1 ---
    ws = wb.active
    ws.title = f"{area}㎡总价测算"
    ws["A1"] = "圣都整装 AEs vs A5s 硬装全案对比测算"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    ws["A2"] = (
        f"房产商售卖面积 {sales_area}㎡ · 计价面积 {area}㎡ · "
        "落在 50≤面积<80 档 · 含15㎡全屋定制 · "
        "「全案基础价」不含管理费（来自宣传页公式）"
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 40

    ws["A4"] = "计价明细"
    ws["A4"].font = Font(bold=True, size=12)
    headers = ["项目", "AEs", "A5s", "差额(A5s-AEs)", "说明"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=5, column=i, value=h)
    style_header(ws, 5, 5)

    rows = [
        ("硬装基础价(50㎡档)", 65000, 75800, None, "宣传页档位起步硬装价"),
        ("超面积单价(元/㎡)", 699, 799, 100, "50≤计价面积<80"),
        ("超出面积(㎡)", round(area - 50, 2), round(area - 50, 2), 0, f"{area}-50"),
        ("超面积费用", round((area - 50) * 699, 2), round((area - 50) * 799, 2), None, "超出面积×单价"),
        ("硬装小计", round(q["aes_hard"], 2), round(q["a5s_hard"], 2), None, "硬装基础+超面积"),
        ("定制-颗粒板(15㎡)", q["custom_p"], q["custom_p"], 0, "标配定制量"),
        ("定制-实木芯(15㎡)", q["custom_s"], q["custom_s"], 0, "相对颗粒板+3000"),
        ("全案基础价(颗粒板)", round(q["aes_base_p"], 2), round(q["a5s_base_p"], 2), None, "不含管理费"),
        ("全案基础价(实木芯)", round(q["aes_base_s"], 2), round(q["a5s_base_s"], 2), None, "不含管理费"),
        ("工程管理费12%(硬装)", round(q["aes_mgmt"], 2), round(q["a5s_mgmt"], 2), None, "定制免收"),
        ("项目经理费2%(硬装估)", round(q["aes_pm"], 2), round(q["a5s_pm"], 2), None, "签约以合同为准"),
        ("含费预估(颗粒板)", round(q["aes_total_p"], 2), round(q["a5s_total_p"], 2), None, "基础+管理+经理费"),
        ("含费预估(实木芯)", round(q["aes_total_s"], 2), round(q["a5s_total_s"], 2), None, "基础+管理+经理费"),
    ]
    for r_i, (name, aes, a5s, forced, note) in enumerate(rows):
        r = 6 + r_i
        diff = forced if forced is not None else round(a5s - aes, 2)
        ws.cell(row=r, column=1, value=name).border = THIN
        c2 = ws.cell(row=r, column=2, value=aes)
        c2.border = THIN
        c2.fill = AES_FILL
        c2.number_format = "#,##0.00"
        c3 = ws.cell(row=r, column=3, value=a5s)
        c3.border = THIN
        c3.fill = A5S_FILL
        c3.number_format = "#,##0.00"
        c4 = ws.cell(row=r, column=4, value=diff)
        c4.border = THIN
        c4.fill = DIFF_FILL
        c4.number_format = "#,##0.00"
        ws.cell(row=r, column=5, value=note).border = THIN

    for r in (13, 14, 17, 18):
        for c in range(1, 6):
            ws.cell(row=r, column=c).font = Font(bold=True)

    ws["A20"] = "关键结论"
    ws["A20"].font = Font(bold=True, size=12)
    ws["A21"] = (
        f"计价面积 {area}㎡：A5s 颗粒板基础价约比 AEs 贵 "
        f"{round(q['a5s_base_p'] - q['aes_base_p'], 2)} 元；含费后约贵 "
        f"{round(q['a5s_total_p'] - q['aes_total_p'], 2)} 元。"
        " 主材高度重合；A5s 主要多出门牌可选 TATA、防水可选西卡、明确含淋浴房、橱柜/地板可选面更宽。"
    )
    ws["A21"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A21:E21")
    ws.row_dimensions[21].height = 70

    ws["A23"] = "注意"
    ws["A23"].font = Font(bold=True, color="B71C1C")
    ws["A24"] = (
        "1) 不含拆改、个性化设计、现场木作等，按实结算。\n"
        "2) 管理费/项目经理费基数以合同为准。\n"
        "3) 材料以展厅实物为准，品牌会调整。\n"
        f"4) 售卖面积 {sales_area}㎡ ≠ 计价面积 {area}㎡，报价按计价面积。"
    )
    ws["A24"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A24:E24")
    ws.row_dimensions[24].height = 80
    ws["A24"].fill = WARN_FILL
    autosize(ws, [36, 16, 16, 16, 42])

    # --- Sheet 2 ---
    ws2 = wb.create_sheet("价目表(提取)")
    ws2["A1"] = "一房一价查询表（含15㎡定制，不含管理费）"
    ws2["A1"].font = Font(bold=True, size=14)

    ws2["A3"] = "AEs"
    ws2["A3"].fill = AES_FILL
    for i, h in enumerate(["计价面积㎡", "颗粒板", "实木芯", "差额"], 1):
        ws2.cell(row=4, column=i, value=h)
    style_header(ws2, 4, 4)
    for i, (a, p, s) in enumerate(AES_PRICE_TABLE):
        r = 5 + i
        for c, v in enumerate([a, p, s, s - p], 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.fill = AES_FILL

    ws2["A12"] = "A5s"
    ws2["A12"].fill = A5S_FILL
    for i, h in enumerate(["计价面积㎡", "颗粒板", "实木芯", "差额"], 1):
        ws2.cell(row=13, column=i, value=h)
    style_header(ws2, 13, 4)
    for i, (a, p, s) in enumerate(A5S_PRICE_TABLE):
        r = 14 + i
        for c, v in enumerate([a, p, s, s - p], 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.fill = A5S_FILL

    ws2["A30"] = "计价逻辑"
    logic = [
        ["产品", "档位", "硬装基础", "超面积规则", "定制", "管理费", "项目经理费"],
        ["AEs", "50㎡一卫", "65000", "50≤S<80：+699/㎡", "13000/16000", "硬装×12%", "2%"],
        ["AEs", "80㎡一卫", "85000", "80≤S<110：+699/㎡", "13000/16000", "硬装×12%", "2%"],
        ["A5s", "50㎡一卫", "75800", "50≤S<80：+799/㎡", "13000/16000", "硬装×12%", "2%"],
        ["A5s", "80㎡一卫", "97500", "80≤S<110：+799/㎡", "13000/16000", "硬装×12%", "2%"],
        ["A5s", "110㎡两卫", "134500", "110≤S<140：+799/㎡", "13000/16000", "硬装×12%", "2%"],
        ["A5s", "≥140㎡", "S×1208", "按计价面积×1208", "13000/16000", "硬装×12%", "2%"],
    ]
    for i, row in enumerate(logic):
        for j, v in enumerate(row, 1):
            cell = ws2.cell(row=31 + i, column=j, value=v)
            cell.border = THIN
            if i == 0:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
    autosize(ws2, [12, 14, 14, 28, 14, 12, 12])

    # --- Sheet 3 ---
    ws3 = wb.create_sheet("配置对比")
    ws3["A1"] = "硬装主材/辅材配置对比"
    ws3["A1"].font = Font(bold=True, size=14)
    for i, h in enumerate(["空间", "品类", "AEs", "A5s", "差异判断"], 1):
        ws3.cell(row=3, column=i, value=h)
    style_header(ws3, 3, 5)
    for i, row in enumerate(CONFIGS):
        r = 4 + i
        for j, v in enumerate(row, 1):
            cell = ws3.cell(row=r, column=j, value=v)
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if j == 3:
                cell.fill = AES_FILL
            if j == 4:
                cell.fill = A5S_FILL
            if j == 5 and ("优势" in v or "未明确" in v):
                cell.fill = DIFF_FILL
        ws3.row_dimensions[r].height = 30
    autosize(ws3, [10, 22, 42, 42, 28])

    # --- Sheet 4 ---
    ws4 = wb.create_sheet("决策摘要")
    ws4["A1"] = f"怎么选：AEs vs A5s（硬装 · {area}㎡）"
    ws4["A1"].font = Font(bold=True, size=14)
    summary = [
        ["维度", "结论"],
        [
            "预算优先",
            f"选 AEs。颗粒板基础价约 {round(q['aes_base_p']):,}，比 A5s 少约 "
            f"{round(q['a5s_base_p'] - q['aes_base_p']):,}；含费差距约 "
            f"{round(q['a5s_total_p'] - q['aes_total_p']):,}。",
        ],
        [
            "品牌/细节优先",
            "选 A5s。木门可上 TATA、防水可上西卡、卫浴明确含淋浴房、橱柜/地板可选面更宽。",
        ],
        [
            "性价比判断",
            "主材品牌池高度重合；价差主要买在门/线/防水/淋浴房/橱柜可选范围，而非整套换代。",
        ],
        [
            "颗粒板 vs 实木芯",
            "同产品内实木芯仅 +3,000（15㎡内）。往往比升档到 A5s 更划算。",
        ],
        [
            "建议默认路径",
            "预算紧：AEs + 实木芯。在意木门与淋浴房打包：A5s。签约前按计价面积出正式报价核对。",
        ],
        [
            "待门店确认",
            "淋浴房是否计入AEs；管理费基数；15㎡不够时的增项单价；拆改/新风/地暖是否另计。",
        ],
    ]
    for i, row in enumerate(summary):
        for j, v in enumerate(row, 1):
            cell = ws4.cell(row=3 + i, column=j, value=v)
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if i == 0:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
    ws4.column_dimensions["A"].width = 16
    ws4.column_dimensions["B"].width = 88
    for r in range(4, 10):
        ws4.row_dimensions[r].height = 48

    return wb


def main() -> None:
    p = argparse.ArgumentParser(description="AEs vs A5s 硬装对比 → Excel")
    p.add_argument("--aes", type=Path, help="AEs 长图路径（可选，仅作记录）")
    p.add_argument("--a5s", type=Path, help="A5s 长图路径（可选，仅作记录）")
    p.add_argument("--area", type=float, default=76.34, help="计价面积㎡")
    p.add_argument("--sales-area", type=float, default=90.0, help="售卖面积㎡")
    p.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="输出 xlsx 路径",
    )
    args = p.parse_args()
    out = args.output or Path(f"AEs_vs_A5s_硬装对比_{args.area}㎡.xlsx")
    wb = build_workbook(args.area, args.sales_area)
    wb.save(out)
    print(f"Wrote {out.resolve()}")


if __name__ == "__main__":
    main()
